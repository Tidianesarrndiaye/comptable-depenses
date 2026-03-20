from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .models import Entry, EntryCategory, EntryStatus, EntryType, ExpenseSheet, RecurringEntryTemplate


class HomeViewTests(TestCase):
	def test_home_page_renders_successfully(self):
		response = self.client.get(reverse('expenses:home'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Suivi clair des depenses et des gains.')

	def test_entry_list_page_renders_entries(self):
		sheet = ExpenseSheet.objects.create(name='Budget maison')
		Entry.objects.create(
			sheet=sheet,
			title='Prime',
			entry_type=EntryType.INCOME,
			category=EntryCategory.OTHER,
			amount=Decimal('80.00'),
			entry_date='2026-03-21',
		)

		response = self.client.get(reverse('expenses:entry-list'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Prime')

	def test_entry_list_can_filter_by_status_and_category(self):
		sheet = ExpenseSheet.objects.create(name='Budget maison')
		Entry.objects.create(
			sheet=sheet,
			title='Courses',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.FOOD,
			amount=Decimal('50.00'),
			entry_date='2026-03-21',
			status=EntryStatus.VALIDATED,
		)
		Entry.objects.create(
			sheet=sheet,
			title='Prime draft',
			entry_type=EntryType.INCOME,
			category=EntryCategory.OTHER,
			amount=Decimal('75.00'),
			entry_date='2026-03-21',
			status=EntryStatus.DRAFT,
		)

		response = self.client.get(reverse('expenses:entry-list'), {'status': EntryStatus.VALIDATED, 'category': EntryCategory.FOOD})

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Courses')
		self.assertNotContains(response, 'Prime draft')

	def test_home_page_displays_analytics_sections(self):
		sheet = ExpenseSheet.objects.create(name='Budget maison')
		Entry.objects.create(
			sheet=sheet,
			title='Salaire',
			entry_type=EntryType.INCOME,
			category=EntryCategory.OTHER,
			amount=Decimal('1000.00'),
			entry_date='2026-03-10',
			status=EntryStatus.VALIDATED,
		)

		response = self.client.get(reverse('expenses:home'))

		self.assertEqual(response.status_code, 200)
		self.assertContains(response, 'Vue mensuelle')
		self.assertContains(response, 'Categories dominantes')


class ExpenseSheetBalanceTests(TestCase):
	def test_current_balance_uses_only_validated_entries(self):
		sheet = ExpenseSheet.objects.create(name='Budget principal', starting_balance=Decimal('100.00'))

		Entry.objects.create(
			sheet=sheet,
			title='Salaire',
			entry_type=EntryType.INCOME,
			category=EntryCategory.OTHER,
			amount=Decimal('250.00'),
			entry_date='2026-03-20',
			status=EntryStatus.VALIDATED,
		)
		Entry.objects.create(
			sheet=sheet,
			title='Courses',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.FOOD,
			amount=Decimal('40.00'),
			entry_date='2026-03-20',
			status=EntryStatus.VALIDATED,
		)
		Entry.objects.create(
			sheet=sheet,
			title='Loyer planifie',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.UTILITIES,
			amount=Decimal('300.00'),
			entry_date='2026-03-20',
			effective_date='2026-03-28',
			status=EntryStatus.SCHEDULED,
		)

		self.assertEqual(sheet.current_balance, Decimal('310.00'))


class CreationFlowTests(TestCase):
	def test_sheet_can_be_created_from_form(self):
		response = self.client.post(
			reverse('expenses:sheet-create'),
			{
				'name': 'Suivi familial',
				'starting_balance': '150.00',
				'is_active': 'on',
			},
		)

		self.assertEqual(response.status_code, 302)
		self.assertTrue(ExpenseSheet.objects.filter(name='Suivi familial').exists())

	def test_scheduled_entry_can_be_created_from_form(self):
		sheet = ExpenseSheet.objects.create(name='Budget principal')

		response = self.client.post(
			reverse('expenses:entry-create'),
			{
				'sheet': sheet.pk,
				'title': 'Internet',
				'entry_type': EntryType.EXPENSE,
				'category': EntryCategory.UTILITIES,
				'amount': '45.00',
				'entry_date': '2026-03-22',
				'description': 'Abonnement mensuel',
				'status': EntryStatus.SCHEDULED,
				'effective_date': '',
			},
		)

		self.assertEqual(response.status_code, 302)
		entry = Entry.objects.get(title='Internet')
		self.assertEqual(entry.status, EntryStatus.SCHEDULED)
		self.assertEqual(str(entry.effective_date), '2026-03-22')

	def test_recurring_template_generates_scheduled_entry(self):
		sheet = ExpenseSheet.objects.create(name='Budget principal')
		template = RecurringEntryTemplate.objects.create(
			sheet=sheet,
			name='Loyer',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.UTILITIES,
			amount=Decimal('300.00'),
			day_of_month=25,
		)

		entry, created = template.generate_scheduled_entry(reference_date=__import__('datetime').date(2026, 3, 20))

		self.assertTrue(created)
		self.assertEqual(entry.status, EntryStatus.SCHEDULED)
		self.assertEqual(str(entry.effective_date), '2026-03-25')


class ExportTests(TestCase):
	def test_csv_export_respects_filters(self):
		sheet = ExpenseSheet.objects.create(name='Budget export')
		Entry.objects.create(
			sheet=sheet,
			title='Salaire export',
			entry_type=EntryType.INCOME,
			category=EntryCategory.OTHER,
			amount=Decimal('900.00'),
			entry_date='2026-03-25',
			status=EntryStatus.VALIDATED,
		)
		Entry.objects.create(
			sheet=sheet,
			title='Brouillon export',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.EXTRAS,
			amount=Decimal('20.00'),
			entry_date='2026-03-25',
			status=EntryStatus.DRAFT,
		)

		response = self.client.get(reverse('expenses:entry-export-csv'), {'status': EntryStatus.VALIDATED})

		self.assertEqual(response.status_code, 200)
		self.assertEqual(response['Content-Type'], 'text/csv')
		self.assertIn('attachment; filename="mouvements.csv"', response['Content-Disposition'])
		body = response.content.decode('utf-8')
		self.assertIn('Salaire export', body)
		self.assertNotIn('Brouillon export', body)

	def test_excel_export_respects_filters(self):
		sheet = ExpenseSheet.objects.create(name='Budget excel')
		Entry.objects.create(
			sheet=sheet,
			title='Courses excel',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.FOOD,
			amount=Decimal('35.00'),
			entry_date='2026-03-26',
			status=EntryStatus.VALIDATED,
		)
		Entry.objects.create(
			sheet=sheet,
			title='Transport brouillon',
			entry_type=EntryType.EXPENSE,
			category=EntryCategory.TRANSPORT,
			amount=Decimal('15.00'),
			entry_date='2026-03-26',
			status=EntryStatus.DRAFT,
		)

		response = self.client.get(reverse('expenses:entry-export-excel'), {'status': EntryStatus.VALIDATED})

		self.assertEqual(response.status_code, 200)
		self.assertEqual(
			response['Content-Type'],
			'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
		)
		self.assertIn('attachment; filename="mouvements.xlsx"', response['Content-Disposition'])

		workbook = load_workbook(filename=BytesIO(response.content))
		worksheet = workbook.active
		rows = list(worksheet.iter_rows(values_only=True))
		self.assertEqual(rows[0], ('Date', 'Titre', 'Feuille', 'Type', 'Categorie', 'Montant', 'Statut', 'Date effective'))
		values = [row[1] for row in rows[1:]]
		self.assertIn('Courses excel', values)
		self.assertNotIn('Transport brouillon', values)
